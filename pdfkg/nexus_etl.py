"""
ETL Pipeline para procesar archivos Excel de datos de turismo de Andalucía.
Extrae, transforma y carga métricas de turismo a ArangoDB.
"""

import os
import re
import logging
from typing import Dict, List, Optional, Tuple, Any
from datetime import datetime
from pathlib import Path
import openpyxl
from tqdm import tqdm

if __package__:
    from .nexus_db import NexusDB, initialize_nexus_db
    from .nexus_pdf_parser import parse_pdf_filename, extract_metrics_from_pdf
else:  # pragma: no cover - ejecución directa
    from nexus_db import NexusDB, initialize_nexus_db  # type: ignore
    from nexus_pdf_parser import parse_pdf_filename, extract_metrics_from_pdf  # type: ignore

# Configurar logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

# Mapeo de meses en español a números
MESES_MAP = {
    "ene": 1, "feb": 2, "mar": 3, "abr": 4,
    "may": 5, "jun": 6, "jul": 7, "ago": 8,
    "sep": 9, "oct": 10, "nov": 11, "dic": 12
}

# Categorías de turismo
CATEGORIAS_MAP = {
    "01": "total_turistas",
    "02": "espanoles",
    "03": "andaluces",
    "04": "resto_espana",
    "05": "extranjeros",
    "06": "britanicos",
    "07": "alemanes",
    "08": "otros_mercados",
    "09": "litoral",
    "10": "interior",
    "11": "cruceros",
    "12": "ciudad",
    "13": "cultural",
    "14": "almeria",
    "15": "cadiz",
    "16": "cordoba",
    "17": "granada",
    "18": "huelva",
    "19": "jaen",
    "20": "malaga",
    "21": "sevilla"
}

# Tipos de periodo
PERIODO_TIPOS = {
    "DATOS MENSUALES": "mensual",
    "DATOS ACUMULADO AÑO": "acumulado",
    "DATOS AÑO COMPLETO": "anual"
}


SECTION_NAME_MAP = {
    "VIAJEROS EN ESTABLECIMIENTOS HOTELEROS DE ANDALUCÍA": "Número de viajeros en establecimientos hoteleros",
    "PERNOCTACIONES EN ESTABLECIMIENTOS HOTELEROS DE ANDALUCÍA": "Número de pernoctaciones en establecimientos hoteleros",
}


def parse_filename(filename: str) -> Optional[Dict[str, Any]]:
    """
    Parsea el nombre de archivo para extraer metadata.

    Args:
        filename: Nombre del archivo (ej: "01_total_turistas_ene25.xlsx")

    Returns:
        Dict con metadata o None si el formato no es válido
    """
    try:
        # Remover extensión y sufijo "_limpio"
        base_name = filename.replace(".xlsx", "").replace("_limpio", "")

        # Pattern: {numero}_{categoria}_{mes}{año}
        # Ejemplos: "01_total_turistas_ene25", "11_cruceros_feb24"
        pattern = r"^(\d{2})_([a-z_]+)_([a-z]{3})(\d{2})$"
        match = re.match(pattern, base_name)

        if not match:
            logger.warning(f"Nombre de archivo no coincide con el patrón: {filename}")
            return None

        numero, categoria_raw, mes_str, anio_str = match.groups()

        # Validar y convertir mes
        mes = MESES_MAP.get(mes_str.lower())
        if mes is None:
            logger.warning(f"Mes no reconocido: {mes_str}")
            return None

        # Convertir año (25 -> 2025, 24 -> 2024, etc.)
        anio = 2000 + int(anio_str)

        # Obtener nombre de categoría
        categoria = CATEGORIAS_MAP.get(numero, categoria_raw)

        # Determinar si es versión limpia
        es_limpio = "_limpio" in filename

        return {
            "numero": numero,
            "categoria": categoria,
            "mes": mes,
            "mes_str": mes_str,
            "anio": anio,
            "es_limpio": es_limpio,
            "archivo_original": filename
        }

    except Exception as e:
        logger.error(f"Error al parsear nombre de archivo '{filename}': {e}")
        return None


def extract_metrics_from_excel(file_path: str, metadata: Dict) -> List[Dict[str, Any]]:
    """
    Extrae métricas de un archivo Excel.

    Args:
        file_path: Ruta al archivo Excel
        metadata: Metadata extraída del nombre del archivo

    Returns:
        Lista de métricas extraídas
    """
    metrics = []

    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)

        # Seleccionar hoja válida (acepta 'ficha' y alternativas conocidas)
        sheet_candidates = ("ficha", "hojafinal")
        sheet_name = None
        lower_names = {name.lower(): name for name in wb.sheetnames}
        for candidate in sheet_candidates:
            match = lower_names.get(candidate)
            if match:
                sheet_name = match
                break

        if not sheet_name:
            logger.warning(f"Sheet 'ficha' ni 'HojaFinal' encontrada en {file_path}")
            return metrics

        ws = wb[sheet_name]

        def to_float(value):
            if value is None:
                return None
            if isinstance(value, (int, float)):
                return float(value)
            if isinstance(value, str):
                candidate = value.strip()
                if not candidate:
                    return None
                candidate = candidate.replace(' ', '').replace(' ', '')
                if candidate.count(',') == 1 and candidate.count('.') > 1:
                    candidate = candidate.replace('.', '').replace(',', '.')
                else:
                    candidate = candidate.replace(',', '.')
                try:
                    return float(candidate)
                except ValueError:
                    return None
            return None

        period_configs = []
        current_section = None

        def infer_period_type(periodo_str: str) -> str:
            if not periodo_str:
                return "mensual"
            periodo_lower = periodo_str.lower()
            if "año completo" in periodo_lower:
                return "anual"
            if "acumulado" in periodo_lower:
                return "acumulado"
            # Detectar rangos tipo 'enero-abril'
            first_segment = periodo_lower.split(" - ")[0]
            if "-" in first_segment and any(mes in first_segment for mes in MESES_MAP.keys()):
                return "acumulado"
            return "mensual"

        for row_idx, row in enumerate(ws.iter_rows(min_row=10, max_row=200), start=10):
            try:
                values = [cell.value for cell in row]

                if not any(values):
                    continue

                # Detectar filas que definen configuraciones de periodo (ej. DATOS MENSUALES)
                configs_found = []
                for idx, value in enumerate(values):
                    if isinstance(value, str):
                        key = value.strip().upper()
                        if key in PERIODO_TIPOS:
                            configs_found.append((PERIODO_TIPOS[key], idx))
                if configs_found:
                    period_configs = []
                    for periodo, idx_base in configs_found:
                        period_configs.append({
                            "periodo_tipo": periodo,
                            "valor_idx": idx_base,
                            "variacion_idx": idx_base + 1,
                            "descripcion_idx": idx_base + 2,
                        })
                    continue

                col_b = values[1] if len(values) > 1 else None

                # Detectar encabezados de secciones (ej. VIAJEROS..., PERNOCTACIONES...)
                if isinstance(col_b, str):
                    col_b_clean = col_b.strip()
                    upper_col_b = col_b_clean.upper()
                    has_numeric = any(isinstance(v, (int, float)) or isinstance(v, str) and to_float(v) is not None for v in values[2:5])
                    if not has_numeric and col_b_clean and upper_col_b not in PERIODO_TIPOS:
                        current_section = SECTION_NAME_MAP.get(upper_col_b, col_b_clean)
                        continue

                if not isinstance(col_b, str):
                    continue

                metrica_base = col_b.strip()
                if not metrica_base or metrica_base.upper() in PERIODO_TIPOS:
                    continue

                configs_for_row = period_configs.copy() if period_configs else []
                if not configs_for_row and len(values) > 2:
                    numeric_found_idx = None
                    for idx_candidate in range(2, len(values)):
                        candidate_val = to_float(values[idx_candidate])
                        if candidate_val is not None and abs(candidate_val) >= 1:
                            numeric_found_idx = idx_candidate
                            break

                    if numeric_found_idx is not None:
                        periodo_str = ""
                        for idx_candidate in range(numeric_found_idx + 1, len(values)):
                            candidate_period = values[idx_candidate]
                            if isinstance(candidate_period, str) and candidate_period.strip():
                                periodo_str = candidate_period
                                break

                        configs_for_row = [{
                            "periodo_tipo": infer_period_type(periodo_str),
                            "valor_idx": numeric_found_idx,
                            "variacion_idx": numeric_found_idx + 1,
                            "descripcion_idx": numeric_found_idx + 2,
                        }]

                if not configs_for_row:
                    continue

                for config in configs_for_row:
                    valor_idx = config["valor_idx"]
                    variacion_idx = config["variacion_idx"]
                    descripcion_idx = config.get("descripcion_idx", valor_idx + 2)

                    if valor_idx >= len(values):
                        continue

                    metrica_valor = values[valor_idx]
                    metrica_valor_num = to_float(metrica_valor)

                    if metrica_valor_num is None or (isinstance(metrica_valor_num, float) and abs(metrica_valor_num) < 1):
                        shifted = False
                        for shift in range(1, 4):
                            candidate_idx = valor_idx + shift
                            if candidate_idx >= len(values):
                                break
                            candidate_value = to_float(values[candidate_idx])
                            if candidate_value is not None and abs(candidate_value) >= 1:
                                valor_idx = candidate_idx
                                metrica_valor_num = candidate_value
                                shifted = True
                                break

                        if not shifted:
                            for shift in range(1, valor_idx):
                                candidate_idx = valor_idx - shift
                                if candidate_idx <= 1:
                                    break
                                candidate_value = to_float(values[candidate_idx])
                                if candidate_value is not None and abs(candidate_value) >= 1:
                                    valor_idx = candidate_idx
                                    metrica_valor_num = candidate_value
                                    shifted = True
                                    break

                    if metrica_valor_num is None or (isinstance(metrica_valor_num, float) and abs(metrica_valor_num) < 1):
                        continue

                    variacion_interanual = None
                    if variacion_idx < len(values):
                        variacion_interanual = to_float(values[variacion_idx])

                    if variacion_interanual is None or (isinstance(variacion_interanual, float) and abs(variacion_interanual) >= abs(metrica_valor_num)):
                        for candidate_idx in range(valor_idx + 1, len(values)):
                            candidate_value = to_float(values[candidate_idx])
                            if candidate_value is not None and abs(candidate_value) < abs(metrica_valor_num):
                                variacion_interanual = candidate_value
                                break

                    periodo_str = ""
                    if descripcion_idx < len(values):
                        periodo_val = values[descripcion_idx]
                        periodo_str = str(periodo_val) if periodo_val else ""

                    if not periodo_str:
                        for candidate_idx in range(valor_idx + 1, len(values)):
                            candidate_period = values[candidate_idx]
                            if isinstance(candidate_period, str) and candidate_period.strip():
                                periodo_str = candidate_period
                                break

                    seccion_normalizada = current_section or SECTION_NAME_MAP.get(metrica_base.upper())
                    if seccion_normalizada:
                        metrica_nombre = f"{seccion_normalizada} - {metrica_base}"
                    else:
                        metrica_nombre = metrica_base

                    metric_doc = {
                        "categoria": metadata["categoria"],
                        "mes": metadata["mes"],
                        "mes_str": metadata["mes_str"],
                        "anio": metadata["anio"],
                        "periodo_tipo": config["periodo_tipo"],
                        "metrica_nombre": metrica_nombre,
                        "metrica_valor": metrica_valor_num,
                        "metrica_valor_raw": str(metrica_valor) if metrica_valor is not None else None,
                        "variacion_interanual": variacion_interanual,
                        "periodo_descripcion": periodo_str,
                        "fuente_archivo": metadata["archivo_original"],
                        "es_limpio": metadata["es_limpio"],
                        "timestamp_ingestion": datetime.now().isoformat(),
                        "fila_excel": row_idx,
                    }

                    provincias = ["almeria", "cadiz", "cordoba", "granada", "huelva", "jaen", "malaga", "sevilla"]
                    if metadata["categoria"] in provincias:
                        metric_doc["provincia"] = metadata["categoria"].capitalize()

                    metrics.append(metric_doc)

            except Exception as e:
                logger.warning(f"Error al procesar fila {row_idx} en {file_path}: {e}")
                continue

        wb.close()
        logger.info(f"Extraídas {len(metrics)} métricas de {file_path}")

    except Exception as e:
        logger.error(f"Error al procesar archivo Excel {file_path}: {e}")

    return metrics




def process_excel_file(file_path: str, db: NexusDB) -> Tuple[int, int]:
    """
    Procesa un archivo Excel individual y carga sus métricas a la DB.

    Args:
        file_path: Ruta al archivo Excel
        db: Instancia de NexusDB

    Returns:
        Tupla (métricas_extraídas, métricas_cargadas)
    """
    filename = os.path.basename(file_path)

    # Parsear metadata del nombre de archivo
    metadata = parse_filename(filename)
    if metadata is None:
        logger.warning(f"Saltando archivo con nombre inválido: {filename}")
        return 0, 0

    # Extraer métricas del Excel
    metrics = extract_metrics_from_excel(file_path, metadata)

    if not metrics:
        logger.warning(f"No se extrajeron métricas de {filename}")
        return 0, 0

    # Cargar métricas a la DB
    inserted = db.bulk_insert_metrics(metrics)

    return len(metrics), inserted


def process_pdf_file(file_path: str, db: NexusDB) -> Tuple[int, int]:
    """
    Procesa un archivo PDF individual y carga sus métricas a la DB.

    Args:
        file_path: Ruta al archivo PDF
        db: Instancia de NexusDB

    Returns:
        Tupla (métricas_extraídas, métricas_cargadas)
    """
    filename = os.path.basename(file_path)

    # Parsear metadata del nombre de archivo
    metadata = parse_pdf_filename(filename)
    if metadata is None:
        logger.warning(f"Saltando archivo PDF con nombre inválido: {filename}")
        return 0, 0

    # Extraer métricas del PDF
    metrics = extract_metrics_from_pdf(file_path, metadata)

    if not metrics:
        logger.warning(f"No se extrajeron métricas de {filename}")
        return 0, 0

    # Cargar métricas a la DB
    inserted = db.bulk_insert_metrics(metrics)

    return len(metrics), inserted


def run_etl(
    directory: str = "nexus",
    category_filter: Optional[str] = None,
    year_filter: Optional[int] = None,
    month_filter: Optional[int] = None,
    clear_before: bool = False
) -> Dict[str, Any]:
    """
    Ejecuta el pipeline ETL completo.

    Args:
        directory: Directorio con archivos Excel
        category_filter: Filtrar por categoría específica (ej: "cruceros")
        year_filter: Filtrar por año específico (ej: 2025)
        month_filter: Filtrar por mes específico (1-12)
        clear_before: Si True, limpia la colección antes de cargar

    Returns:
        Dict con estadísticas del proceso
    """
    logger.info("=" * 80)
    logger.info("INICIANDO ETL PIPELINE - NEXUS TURISMO ANDALUCÍA")
    logger.info("=" * 80)

    start_time = datetime.now()

    # Conectar a la base de datos
    db = initialize_nexus_db()
    if db is None:
        logger.error("No se pudo inicializar la base de datos")
        return {"error": "Database initialization failed"}

    # Limpiar colección si se solicita
    if clear_before:
        logger.warning("Limpiando colección antes de cargar datos...")
        db.clear_collection("metricas_turismo")

    # Buscar archivos Excel
    excel_files = list(Path(directory).glob("*.xlsx"))
    logger.info(f"Encontrados {len(excel_files)} archivos Excel en '{directory}'")

    # Filtrar archivos según criterios
    filtered_files = []
    for file_path in excel_files:
        filename = file_path.name
        metadata = parse_filename(filename)

        if metadata is None:
            continue

        # Aplicar filtros
        if category_filter and metadata["categoria"] != category_filter:
            continue
        if year_filter and metadata["anio"] != year_filter:
            continue
        if month_filter and metadata["mes"] != month_filter:
            continue

        filtered_files.append(file_path)

    logger.info(f"Procesando {len(filtered_files)} archivos después de aplicar filtros")

    # Procesar archivos
    total_extracted = 0
    total_loaded = 0
    files_processed = 0
    files_failed = 0

    for file_path in tqdm(filtered_files, desc="Procesando archivos"):
        try:
            extracted, loaded = process_excel_file(str(file_path), db)
            total_extracted += extracted
            total_loaded += loaded
            files_processed += 1
        except Exception as e:
            logger.error(f"Error al procesar {file_path.name}: {e}")
            files_failed += 1

    # Procesar archivos PDF
    logger.info("=" * 80)
    logger.info("Procesando archivos PDF...")
    logger.info("=" * 80)

    pdf_files = list(Path(directory).glob("*.pdf"))
    logger.info(f"Encontrados {len(pdf_files)} archivos PDF en '{directory}'")

    # Filtrar PDFs según criterios
    filtered_pdfs = []
    for file_path in pdf_files:
        filename = file_path.name
        metadata = parse_pdf_filename(filename)

        if metadata is None:
            continue

        # Aplicar filtros
        if year_filter and metadata["anio"] != year_filter:
            continue
        if month_filter and metadata["mes"] != month_filter:
            continue

        filtered_pdfs.append(file_path)

    logger.info(f"Procesando {len(filtered_pdfs)} archivos PDF después de aplicar filtros")

    pdf_extracted = 0
    pdf_loaded = 0
    pdf_processed = 0
    pdf_failed = 0

    for file_path in tqdm(filtered_pdfs, desc="Procesando PDFs"):
        try:
            extracted, loaded = process_pdf_file(str(file_path), db)
            pdf_extracted += extracted
            pdf_loaded += loaded
            if extracted > 0:
                pdf_processed += 1
        except Exception as e:
            logger.error(f"Error al procesar PDF {file_path.name}: {e}")
            pdf_failed += 1

    # Actualizar totales
    total_extracted += pdf_extracted
    total_loaded += pdf_loaded
    files_processed += pdf_processed
    files_failed += pdf_failed

    # Calcular estadísticas finales
    end_time = datetime.now()
    duration = (end_time - start_time).total_seconds()

    stats = {
        "archivos_encontrados": len(excel_files),
        "archivos_filtrados": len(filtered_files),
        "pdfs_encontrados": len(pdf_files),
        "pdfs_filtrados": len(filtered_pdfs),
        "archivos_procesados": files_processed,
        "archivos_fallidos": files_failed,
        "excel_procesados": files_processed - pdf_processed,
        "pdf_procesados": pdf_processed,
        "metricas_extraidas": total_extracted,
        "metricas_cargadas": total_loaded,
        "duracion_segundos": duration,
        "timestamp": end_time.isoformat()
    }

    # Obtener estadísticas de la DB
    db_stats = db.get_stats()
    stats["estadisticas_db"] = db_stats

    # Cerrar conexión
    db.close()

    # Mostrar resumen
    logger.info("=" * 80)
    logger.info("RESUMEN ETL")
    logger.info("=" * 80)
    logger.info(f"📁 Excel procesados: {files_processed - pdf_processed}/{len(filtered_files)}")
    logger.info(f"📄 PDFs procesados: {pdf_processed}/{len(filtered_pdfs)}")
    logger.info(f"✅ Total archivos procesados: {files_processed}")
    logger.info(f"❌ Archivos fallidos: {files_failed}")
    logger.info(f"📊 Métricas extraídas: {total_extracted}")
    logger.info(f"💾 Métricas cargadas: {total_loaded}")
    logger.info(f"⏱️  Duración: {duration:.2f} segundos")
    logger.info(f"🗄️  Total en DB: {db_stats.get('total_metricas', 0)} métricas")
    logger.info("=" * 80)

    return stats


if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser(description="ETL Pipeline para datos de turismo")
    parser.add_argument("--directory", "-d", default="./nexus/data/ultimos_datos_turisticos", help="Directorio con archivos Excel")
    parser.add_argument("--category", "-c", help="Filtrar por categoría")
    parser.add_argument("--year", "-y", type=int, help="Filtrar por año")
    parser.add_argument("--month", "-m", type=int, help="Filtrar por mes (1-12)")
    parser.add_argument("--clear", action="store_true", help="Limpiar colección antes de cargar")

    args = parser.parse_args()

    # Ejecutar ETL
    stats = run_etl(
        directory=args.directory,
        category_filter=args.category,
        year_filter=args.year,
        month_filter=args.month,
        clear_before=args.clear
    )

    if "error" not in stats:
        print(f"\n✅ ETL completado exitosamente!")
        print(f"📊 {stats['metricas_cargadas']} métricas cargadas en {stats['duracion_segundos']:.2f}s")
    else:
        print(f"\n❌ ETL falló: {stats['error']}")
