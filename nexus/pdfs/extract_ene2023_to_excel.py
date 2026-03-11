
#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Extrae métricas de "DATOS BÁSICOS DEL TURISMO EN ANDALUCÍA" (ENERO 2023) del PDF
y actualiza la hoja "HojaFinal" de una planilla Excel (plantilla 01_total_turistas_*.xlsx).
Si faltan algunas métricas, puede intentar un fallback con un LLM vía vLLM.

Uso:
  python extract_ene2023_to_excel.py \
      --pdf /path/ultimos-datos_ene23.pdf \
      --excel /path/01_total_turistas_jun25.xlsx \
      --out /path/01_total_turistas_ene23.xlsx

Variables de entorno opcionales:
  VLLM_BASE_URL (default: http://181.66.252.169:8000/v1)
  VLLM_MODEL     (default: Qwen3-8B-AWQ)
  USE_VLLM       (default: 1)  # 0 para deshabilitar fallback LLM
"""
import os
import re
import json
import argparse
import unicodedata

def extract_pdf_text(pdf_path: str, pdfminer=None) -> str:
    # 1) PyPDF2
    try:
        import PyPDF2
        txt = []
        with open(pdf_path, "rb") as f:
            reader = PyPDF2.PdfReader(f)
            for p in reader.pages:
                txt.append(p.extract_text() or "")
        return "\n".join(txt)
    except Exception as e1:
        # 2) pdfminer.six
        try:
            from pdfminer.high_level import extract_text as pdfminer_extract_text
            return pdfminer_extract_text(pdf_path)
        except Exception as e2:
            raise RuntimeError(f"No se pudo extraer texto del PDF con PyPDF2 ni pdfminer: {e1} | {e2}")

def normalize_text(t: str) -> str:
    # Limpieza de caracteres raros y diacríticos; colapsar espacios, a minúsculas
    t = t.replace("¯", "n").replace("¸", "u").replace("«", "i").replace("·", ".")
    t = unicodedata.normalize("NFD", t)
    t = "".join(ch for ch in t if not unicodedata.combining(ch))
    t = t.lower()
    t = re.sub(r"\s+", " ", t)
    return t

def parse_eu_number(s: str, percent: bool=False):
    s = s.strip()
    m = re.search(r"(\d{1,3}(?:\.\d{3})*(?:,\d+)?|\d+(?:,\d+)?)(%)?", s)
    if not m: 
        return None
    num = m.group(1).replace(".", "").replace(",", ".")
    val = float(num)
    if percent or m.group(2) == "%":
        val /= 100.0
    return val

def extract_monthly_values(norm_text: str):
    # Buscamos las tres primeras métricas mensuales justo antes de "enero - 2023":
    # 1) viajeros, 2) pernoctaciones, 3) cuota (%)
    positions = [m.start() for m in re.finditer(r"enero\s*-\s*2023", norm_text)]
    vals_seq = []
    for pos in positions[:6]:
        window = norm_text[max(0,pos-120):pos]
        mnum = re.findall(r"(\d{1,3}(?:\.\d{3})*(?:,\d+)?|\d+(?:,\d+)?)(\s*%)?", window)
        vals_seq.append(mnum[-2] if len(mnum) >= 2 else None)

    viajeros_val = parse_eu_number(vals_seq[0][0]) if vals_seq and vals_seq[0] else None
    pernoct_val = parse_eu_number(vals_seq[2][0]) if len(vals_seq)>=3 and vals_seq[2] else None
    cuota_val    = parse_eu_number(vals_seq[4][0], percent=True) if len(vals_seq)>=5 and vals_seq[4] else None

    # Llegadas (suele venir como "-" en ENE-2023); si aparece número al lado lo tomamos, si no -> None
    llegadas = None
    m_lleg = re.search(r"llegadas de pasajeros a aeropuertos andaluces\s+([0-9\.\-,]+)\s+", norm_text)
    if m_lleg:
        token = m_lleg.group(1).strip()
        if token != "-":
            llegadas = parse_eu_number(token)

    return viajeros_val, pernoct_val, cuota_val, llegadas

def extract_q4_values(norm_text: str):
    # Métricas que aparecen como 4º trim 2022 en el boletín de ENE-2023
    def get_one(pat):
        m = re.search(pat, norm_text)
        return parse_eu_number(m.group(1)) if m else None

    turistas_mill = get_one(r"numero de turistas\s*\(millones\)\s+([0-9\.,]+)\s+[0-9\.,]+\s*%\s*4")
    if turistas_mill is None:
        turistas_mill = get_one(r"numero de turistas\s+([0-9\.,]+)\s+[0-9\.,]+\s*%\s*4")

    estancia = get_one(r"estancia media\s*\(numero de dias\)\s*([0-9\.,]+)\s*[-\+][0-9\.,]*\s*4")
    gasto    = get_one(r"gasto medio diario\s*\(euros\)\s*([0-9\.,]+)\s*[0-9\.,]*\s*4")
    return turistas_mill, estancia, gasto

def call_vllm_fallback(text: str):
    """Intenta estructurar las métricas con un LLM vLLM si faltan campos."""
    base = os.getenv("VLLM_BASE_URL", "http://181.66.252.169:8000/v1")
    model = os.getenv("VLLM_MODEL", "Qwen3-8B-AWQ")
    use   = os.getenv("USE_VLLM", "1") != "0"
    if not use:
        return None

    prompt = f"""
Eres un asistente que extrae datos de un boletín de turismo de Andalucía (enero 2023).
Del siguiente texto, devuelve un JSON con estas claves (usa null si falta):
{{
  "numero_viajeros": int,
  "numero_pernoctaciones": int,
  "cuota_pernoct_espana": float,      // proporción en [0,1], NO porcentaje
  "llegadas_aeropuertos": int|null,
  "numero_turistas_millones": float,
  "estancia_media_dias": float,
  "gasto_medio_diario_eur": float
}}
Texto:
{text}
Responde SOLO el JSON.
"""
    try:
        import requests
        payload = {
            "model": model,
            "messages": [{"role": "user", "content": prompt}],
            "temperature": 0.0,
        }
        resp = requests.post(f"{base}/chat/completions", json=payload, timeout=60)
        resp.raise_for_status()
        content = resp.json()["choices"][0]["message"]["content"]
        # Intentar parsear JSON
        start = content.find("{"); end = content.rfind("}") + 1
        if start != -1 and end != -1:
            data = json.loads(content[start:end])
            return data
    except Exception as e:
        # Si falla, devolvemos None
        return None

def update_excel(excel_path: str, out_path: str, values: dict):
    from openpyxl import load_workbook
    wb = load_workbook(excel_path)
    ws = wb["HojaFinal"]
    # Mapear fila por etiqueta en col B
    label_to_row = {}
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=2, max_col=2):
        cell = row[0]
        if isinstance(cell.value, str):
            label_to_row[cell.value.strip()] = cell.row

    for label, val in values.items():
        row = label_to_row.get(label)
        if row:
            ws.cell(row=row, column=3).value = val
    wb.save(out_path)

def main(pdf_fname):
    raw = extract_pdf_text(pdf_fname)
    norm = normalize_text(raw)

    # 1) Extraer métricas mensuales (enero-2023): viajeros, pernoctaciones, cuota, llegadas
    viajeros, pernoct, cuota, llegadas = extract_monthly_values(norm)
    # 2) Extraer métricas de 4º Trim 2022: numero turistas (millones), estancia, gasto medio
    turistas_mill, estancia, gasto = extract_q4_values(norm)

    values = {
        "Número de viajeros en establecimientos hoteleros": viajeros,
        "Número de pernoctaciones en establecimientos hoteleros": pernoct,
        "Cuota (% sobre total pernoctaciones en España)": cuota,
        "Llegadas de pasajeros a aeropuertos andaluces": llegadas,
        "Número de turistas (millones)": turistas_mill,
        "Estancia Media (número de días)": estancia,
        "Gasto medio diario (euros)": gasto,
    }

    # Si algo crítico falta, intentar LLM fallback
    needs_fallback = any(values[k] is None for k in [
        "Número de viajeros en establecimientos hoteleros",
        "Número de pernoctaciones en establecimientos hoteleros",
        "Cuota (% sobre total pernoctaciones en España)"
    ])
    if needs_fallback:
        llm = call_vllm_fallback(norm)
        if llm:
            # Merge manteniendo los ya extraídos
            values.setdefault("Número de viajeros en establecimientos hoteleros", llm.get("numero_viajeros"))
            values.setdefault("Número de pernoctaciones en establecimientos hoteleros", llm.get("numero_pernoctaciones"))
            values.setdefault("Cuota (% sobre total pernoctaciones en España)", llm.get("cuota_pernoct_espana"))
            values.setdefault("Llegadas de pasajeros a aeropuertos andaluces", llm.get("llegadas_aeropuertos"))
            values.setdefault("Número de turistas (millones)", llm.get("numero_turistas_millones"))
            values.setdefault("Estancia Media (número de días)", llm.get("estancia_media_dias"))
            values.setdefault("Gasto medio diario (euros)", llm.get("gasto_medio_diario_eur"))

    # Escribir Excel
    # update_excel(args.excel, args.out, values)

    # Reporte en consola
    print("Valores extraídos/escritos (ENE-2023):")
    for k, v in values.items():
        print(f" - {k}: {v}")

if __name__ == "__main__":
    main(pdf_fname='/home/maxim/Downloads/ultimos-datos_ene23.pdf')
