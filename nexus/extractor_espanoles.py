import pandas as pd
import psycopg2
from psycopg2.extras import execute_values
import openpyxl
import re
from datetime import datetime
import logging
from pathlib import Path
import os

# Configurar logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)


class EspanolesDataExtractor:
    def __init__(self, excel_file, db_config):
        self.excel_file = excel_file
        self.db_config = db_config
        self.connection = None

    def conectar_db(self):
        """Conectar a la base de datos PostgreSQL"""
        try:
            self.connection = psycopg2.connect(**self.db_config)
            logger.info("Conexión a base de datos establecida")
        except Exception as e:
            logger.error(f"Error conectando a la base de datos: {e}")
            raise

    def extraer_fecha_del_archivo(self, sheet):
        """Extraer año y mes, preferentemente desde el nombre del archivo"""
        nombre = Path(self.excel_file).name.lower()

        meses_abbr = {
            'ene': 1, 'feb': 2, 'mar': 3, 'abr': 4, 'may': 5, 'jun': 6,
            'jul': 7, 'ago': 8, 'sep': 9, 'oct': 10, 'nov': 11, 'dic': 12,
        }
        meses_completos = {
            'enero': 1, 'febrero': 2, 'marzo': 3, 'abril': 4, 'mayo': 5, 'junio': 6,
            'julio': 7, 'agosto': 8, 'septiembre': 9, 'octubre': 10, 'noviembre': 11, 'diciembre': 12,
        }

        # 1) Intentar: abrev + (2|4 dígitos)
        m = re.search(r"(ene|feb|mar|abr|may|jun|jul|ago|sep|oct|nov|dic)[_-]?(\d{2}|\d{4})", nombre)
        if m:
            mes_txt, year_txt = m.groups()
            mes = meses_abbr[mes_txt]
            año = int(year_txt)
            if año < 100:
                año += 2000
            logger.info(f"Fecha extraída desde nombre: {mes}/{año}")
            return año, mes

        # 2) Intentar: mes completo + (2|4 dígitos)
        m = re.search(
            r"(enero|febrero|marzo|abril|mayo|junio|julio|agosto|septiembre|octubre|noviembre|diciembre)[_-]?(\d{2}|\d{4})",
            nombre)
        if m:
            mes_txt, year_txt = m.groups()
            mes = meses_completos[mes_txt]
            año = int(year_txt)
            if año < 100:
                año += 2000
            logger.info(f"Fecha extraída desde nombre: {mes}/{año}")
            return año, mes

        # Fallback: detectar en la hoja buscando "junio - 2025" etc.
        año_detectado = None
        mes_detectado = None
        for i in range(min(30, sheet.max_row)):
            for j in range(min(10, sheet.max_column)):
                cell_value = sheet.cell(row=i + 1, column=j + 1).value
                if isinstance(cell_value, str):
                    texto = cell_value.lower()
                    # Buscar patrones como "junio - 2025"
                    for mes_txt, mes_num in meses_completos.items():
                        if mes_txt in texto and " - " in texto:
                            y = re.search(r"(20\d{2})", texto)
                            if y:
                                año_detectado = int(y.group(1))
                                mes_detectado = mes_num
                                logger.info(f"Fecha extraída desde contenido: {mes_detectado}/{año_detectado}")
                                return año_detectado, mes_detectado

        # Si no se detecta nada, usar defaults
        logger.warning("No se pudo extraer fecha; se usará junio 2025 por defecto")
        return 2025, 6

    def leer_excel_y_extraer_datos(self):
        """Leer el archivo Excel y extraer los datos de españoles"""
        logger.info(f"Leyendo archivo Excel: {self.excel_file}")

        try:
            # Leer con openpyxl para mayor control
            workbook = openpyxl.load_workbook(self.excel_file)
            sheet = workbook.active

            año, mes = self.extraer_fecha_del_archivo(sheet)

            viajeros_hoteles = None
            pernoctaciones_hoteles = None
            viajeros_ya_encontrado = False
            pernoctaciones_ya_encontradas = False

            # Iterar por todas las filas buscando SOLO los primeros datos mensuales
            for row_num in range(1, sheet.max_row + 1):
                row_values = []
                for col_num in range(1, min(6, sheet.max_column + 1)):
                    cell_value = sheet.cell(row=row_num, column=col_num).value
                    row_values.append(cell_value)

                # Verificar si esta fila contiene datos mensuales (no acumulados)
                if len(row_values) >= 5 and row_values[1] is not None:
                    texto_descripcion = str(row_values[1]).lower()
                    periodo_texto = str(row_values[4] or '').lower()

                    # Buscar datos mensuales puros (no acumulados ni anuales)
                    es_mensual = (
                            'enero-' not in periodo_texto and
                            'año' not in periodo_texto and
                            'acumulado' not in periodo_texto and
                            '-' in periodo_texto and
                            str(año) in periodo_texto and
                            len(periodo_texto.split('-')) == 2  # Solo "junio - 2025", no "enero-junio - 2025"
                    )

                    if es_mensual:
                        if ('viajeros' in texto_descripcion and 'hoteleros' in texto_descripcion and
                                not viajeros_ya_encontrado):
                            viajeros_hoteles = self._convertir_a_numero(row_values[2])
                            viajeros_ya_encontrado = True
                            logger.info(
                                f"Viajeros MENSUALES encontrados en fila {row_num}: {viajeros_hoteles} - Período: {row_values[4]}")

                        elif ('pernoctaciones' in texto_descripcion and 'hoteleros' in texto_descripcion and
                              not pernoctaciones_ya_encontradas):
                            pernoctaciones_hoteles = self._convertir_a_numero(row_values[2])
                            pernoctaciones_ya_encontradas = True
                            logger.info(
                                f"Pernoctaciones MENSUALES encontradas en fila {row_num}: {pernoctaciones_hoteles} - Período: {row_values[4]}")

                    # Parar si ya encontramos ambos datos mensuales
                    if viajeros_ya_encontrado and pernoctaciones_ya_encontradas:
                        break

            if not (viajeros_ya_encontrado or pernoctaciones_ya_encontradas):
                logger.warning("No se encontraron datos mensuales de viajeros/pernoctaciones")
                return []

            # Crear el registro para España
            registro_espana = {
                'año': año,
                'mes': mes,
                'codigo_pais': 'ESP',
                'nombre_pais': 'España',
                'viajeros_hoteles': viajeros_hoteles,
                'pernoctaciones_hoteles': pernoctaciones_hoteles
            }

            logger.info(
                f"Datos extraídos para España: Viajeros={viajeros_hoteles}, Pernoctaciones={pernoctaciones_hoteles}")
            return [registro_espana]

        except Exception as e:
            logger.error(f"Error leyendo Excel: {e}")
            raise

    def _convertir_a_numero(self, valor):
        """Convertir valor a número, manejando diferentes formatos"""
        if valor is None:
            return None

        if isinstance(valor, (int, float)):
            return int(valor) if isinstance(valor, float) and valor.is_integer() else int(valor)

        if isinstance(valor, str):
            # Limpiar el string
            valor_limpio = valor.replace(',', '').replace('.', '').strip()
            try:
                return int(valor_limpio)
            except ValueError:
                try:
                    return int(float(valor))
                except ValueError:
                    return None

        return None

    def insertar_datos_en_db(self, datos):
        """Insertar los datos extraídos en la base de datos"""
        if not datos:
            logger.warning("No hay datos para insertar")
            return

        try:
            cursor = self.connection.cursor()

            # Filtrar: solo registros que tengan al menos uno de los 2 campos
            datos_filtrados = [
                d for d in datos
                if (d.get('viajeros_hoteles') is not None or d.get('pernoctaciones_hoteles') is not None)
            ]

            if not datos_filtrados:
                logger.warning("No hay registros válidos (viajeros/pernoctaciones) para insertar")
                return

            # Limpiar datos existentes de España para el mismo período
            año_mes_pairs = set((d['año'], d['mes']) for d in datos_filtrados)

            for año, mes in año_mes_pairs:
                cursor.execute(
                    "DELETE FROM turismo_paises WHERE año = %s AND mes = %s AND codigo_pais = 'ESP'",
                    (año, mes)
                )
                logger.info(f"Eliminados datos existentes de España para {mes}/{año}")

            # Preparar datos para inserción
            valores_insercion = []
            for dato in datos_filtrados:
                valores_insercion.append((
                    dato['año'],
                    dato['mes'],
                    dato['codigo_pais'],
                    dato['nombre_pais'],
                    dato['viajeros_hoteles'],
                    dato['pernoctaciones_hoteles']
                ))

            # Insertar datos
            query_insercion = """
                INSERT INTO turismo_paises 
                (año, mes, codigo_pais, nombre_pais, viajeros_hoteles, pernoctaciones_hoteles)
                VALUES %s
                ON CONFLICT (año, mes, codigo_pais) DO UPDATE SET
                    nombre_pais = EXCLUDED.nombre_pais,
                    viajeros_hoteles = EXCLUDED.viajeros_hoteles,
                    pernoctaciones_hoteles = EXCLUDED.pernoctaciones_hoteles
            """

            execute_values(cursor, query_insercion, valores_insercion)

            self.connection.commit()
            logger.info(f"Insertados {len(valores_insercion)} registros en la base de datos")

        except Exception as e:
            self.connection.rollback()
            logger.error(f"Error insertando datos: {e}")
            raise
        finally:
            cursor.close()

    def procesar_archivo(self):
        """Proceso principal: extraer datos del Excel e insertarlos en la DB"""
        try:
            self.conectar_db()
            datos = self.leer_excel_y_extraer_datos()
            self.insertar_datos_en_db(datos)
            logger.info("Proceso completado exitosamente")

        except Exception as e:
            logger.error(f"Error en el proceso: {e}")
            raise
        finally:
            if self.connection:
                self.connection.close()
                logger.info("Conexión a la base de datos cerrada")


def main():
    # Configuración de la base de datos
    db_config = {
        'host': 'localhost',
        'database': 'nexus',
        'user': 'maxim',  # Cambiar por tu usuario
        'password': 'diganDar',  # Cambiar por tu contraseña
        'port': 5432
    }

    # Directorio con archivos Excel de españoles
    dir_in = './nexus/data/ultimos_datos_turisticos'

    # Buscar archivos de españoles
    files = [fn for fn in os.listdir(dir_in) if '02_espanoles_' in fn and '.xls' in fn and '~' not in fn]

    if not files:
        # Si no encuentra archivos, procesar el archivo específico
        excel_file = '02_espanoles_jun25.xlsx'
        if os.path.exists(excel_file):
            extractor = EspanolesDataExtractor(excel_file, db_config)
            extractor.procesar_archivo()
        else:
            logger.error(f"No se encontró el archivo {excel_file}")
    else:
        # Procesar todos los archivos encontrados
        for excel_file in files:
            excel_file_fname = os.path.join(dir_in, excel_file)
            logger.info(f"Procesando archivo: {excel_file}")

            extractor = EspanolesDataExtractor(excel_file_fname, db_config)
            extractor.procesar_archivo()


if __name__ == "__main__":
    main()