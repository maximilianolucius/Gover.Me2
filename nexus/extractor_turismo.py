import pandas as pd
import psycopg2
from psycopg2.extras import execute_values
import openpyxl
import re
from datetime import datetime
import logging
from pathlib import Path
import os


# Configurar logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)


class TurismoDataExtractor:
    def __init__(self, excel_file, db_config):
        self.excel_file = excel_file
        self.db_config = db_config
        self.connection = None

        # Mapeo de países a códigos ISO
        self.pais_codigo_map = {
            'Francia': 'FRA',
            'Alemania': 'DEU',
            'Reino Unido': 'GBR',
            'Italia': 'ITA',
            'Portugal': 'PRT',
            'Países Bajos': 'NLD',
            'Bélgica': 'BEL',
            'Suiza': 'CHE',
            'Austria': 'AUT',
            'Dinamarca': 'DNK',
            'Suecia': 'SWE',
            'Noruega': 'NOR',
            'Finlandia': 'FIN',
            'Polonia': 'POL',
            'República Checa': 'CZE',
            'Hungría': 'HUN',
            'Eslovaquia': 'SVK',
            'Irlanda': 'IRL',
            'Luxemburgo': 'LUX',
            'Estados Unidos': 'USA',
            'Canadá': 'CAN',
            'Japón': 'JPN',
            'Australia': 'AUS',
            'Nueva Zelanda': 'NZL',
            'Brasil': 'BRA',
            'Argentina': 'ARG',
            'Chile': 'CHL',
            'México': 'MEX'
        }

    def conectar_db(self):
        """Conectar a la base de datos PostgreSQL"""
        try:
            self.connection = psycopg2.connect(**self.db_config)
            logger.info("Conexión a base de datos establecida")
        except Exception as e:
            logger.error(f"Error conectando a la base de datos: {e}")
            raise

    def extraer_fecha_del_archivo(self, sheet):
        """Extraer año y mes, preferentemente desde el nombre del archivo.

        Reglas soportadas en el nombre del archivo (case-insensitive):
        - Abreviaturas ES: ene|feb|mar|abr|may|jun|jul|ago|sep|oct|nov|dic + (aa o aaaa)
          Ej: "..._jun25.xlsx" -> 2025-06, "...-abr2024.xlsx" -> 2024-04
        - Nombres completos ES: enero, febrero, ... + (aa o aaaa)
        - Formatos numéricos: aaaa[-_]?mm
        Si falla, intenta detectar en las primeras celdas de la hoja.
        """

        nombre = Path(self.excel_file).name.lower()

        meses_abbr = {
            'ene': 1, 'feb': 2, 'mar': 3, 'abr': 4, 'may': 5, 'jun': 6,
            'jul': 7, 'ago': 8, 'sep': 9, 'oct': 10, 'nov': 11, 'dic': 12,
        }
        meses_completos = {
            'enero': 1, 'febrero': 2, 'marzo': 3, 'abril': 4, 'mayo': 5, 'junio': 6,
            'julio': 7, 'agosto': 8, 'septiembre': 9, 'octubre': 10, 'noviembre': 11, 'diciembre': 12,
        }

        # 1) Intentar: abrev + (2|4 dígitos)
        m = re.search(r"(ene|feb|mar|abr|may|jun|jul|ago|sep|oct|nov|dic)[_-]?(\d{2}|\d{4})", nombre)
        if m:
            mes_txt, year_txt = m.groups()
            mes = meses_abbr[mes_txt]
            año = int(year_txt)
            if año < 100:
                año += 2000
            logger.info(f"Fecha extraída desde nombre: {mes}/{año}")
            return año, mes

        # 2) Intentar: mes completo + (2|4 dígitos)
        m = re.search(r"(enero|febrero|marzo|abril|mayo|junio|julio|agosto|septiembre|octubre|noviembre|diciembre)[_-]?(\d{2}|\d{4})", nombre)
        if m:
            mes_txt, year_txt = m.groups()
            mes = meses_completos[mes_txt]
            año = int(year_txt)
            if año < 100:
                año += 2000
            logger.info(f"Fecha extraída desde nombre: {mes}/{año}")
            return año, mes

        # 3) Intentar: aaaa[-_]?mm
        m = re.search(r"(20\d{2})[-_]?([01]\d)", nombre)
        if m:
            año = int(m.group(1))
            mes = int(m.group(2))
            if 1 <= mes <= 12:
                logger.info(f"Fecha extraída desde nombre: {mes}/{año}")
                return año, mes

        # Fallback: detectar en la hoja mirando textos como "junio - 2025"
        año_detectado = None
        mes_detectado = None
        for i in range(min(30, sheet.max_row)):
            for j in range(min(10, sheet.max_column)):
                cell_value = sheet.cell(row=i + 1, column=j + 1).value
                if isinstance(cell_value, str):
                    texto = cell_value.lower()
                    # Mes completo + año de 4 dígitos
                    for mes_txt, mes_num in meses_completos.items():
                        if mes_txt in texto:
                            y = re.search(r"(20\d{2})", texto)
                            if y:
                                año_detectado = int(y.group(1))
                                mes_detectado = mes_num
                                break
                    if mes_detectado and año_detectado:
                        break
            if mes_detectado and año_detectado:
                break

        # Si no se detecta nada, usar defaults prudentes pero registrar advertencia
        if año_detectado is None or mes_detectado is None:
            logger.warning("No se pudo extraer fecha; se usará junio 2025 por defecto")
            año_detectado = 2025
            mes_detectado = 6

        logger.info(f"Fecha extraída: {mes_detectado}/{año_detectado}")
        return año_detectado, mes_detectado

    def leer_excel_y_extraer_datos(self):
        """Leer el archivo Excel y extraer los datos de países"""
        logger.info(f"Leyendo archivo Excel: {self.excel_file}")

        try:
            # Leer con openpyxl para mayor control
            workbook = openpyxl.load_workbook(self.excel_file)
            sheet = workbook.active

            año, mes = self.extraer_fecha_del_archivo(sheet)

            datos_extraidos = []
            seccion_actual = None

            # Iterar por todas las filas
            for row_num in range(1, sheet.max_row + 1):
                row_values = []
                for col_num in range(1, min(10, sheet.max_column + 1)):
                    cell_value = sheet.cell(row=row_num, column=col_num).value
                    row_values.append(cell_value)

                # Identificar secciones
                row_text = ' '.join([str(v) for v in row_values if v is not None])

                if 'VIAJEROS EN ESTABLECIMIENTOS HOTELEROS' in row_text:
                    seccion_actual = 'viajeros'
                    logger.info(f"Encontrada sección de viajeros en fila {row_num}")
                elif 'PERNOCTACIONES EN ESTABLECIMIENTOS HOTELEROS' in row_text:
                    seccion_actual = 'pernoctaciones'
                    logger.info(f"Encontrada sección de pernoctaciones en fila {row_num}")

                # Extraer datos de países
                if (seccion_actual and
                        len(row_values) >= 8 and
                        row_values[1] in self.pais_codigo_map and
                        row_values[3] is not None and
                        str(row_values[3]) != '-' and
                        str(row_values[3]).strip() != ''):

                    try:
                        pais_nombre = row_values[1]
                        pais_codigo = self.pais_codigo_map[pais_nombre]

                        # Dato mensual (columna 4)
                        dato_mensual = self._convertir_a_numero(row_values[3])

                        # Buscar si ya existe este país en los datos
                        registro_existente = None
                        for dato in datos_extraidos:
                            if (dato['año'] == año and
                                    dato['mes'] == mes and
                                    dato['codigo_pais'] == pais_codigo):
                                registro_existente = dato
                                break

                        if registro_existente is None:
                            # Crear nuevo registro
                            nuevo_registro = {
                                'año': año,
                                'mes': mes,
                                'codigo_pais': pais_codigo,
                                'nombre_pais': pais_nombre,
                                'viajeros_hoteles': None,
                                'pernoctaciones_hoteles': None
                            }
                            datos_extraidos.append(nuevo_registro)
                            registro_existente = nuevo_registro

                        # Asignar el dato según la sección
                        if seccion_actual == 'viajeros':
                            registro_existente['viajeros_hoteles'] = dato_mensual
                        elif seccion_actual == 'pernoctaciones':
                            registro_existente['pernoctaciones_hoteles'] = dato_mensual

                        logger.debug(f"Procesado: {pais_nombre} - {seccion_actual}: {dato_mensual}")

                    except Exception as e:
                        logger.warning(f"Error procesando fila {row_num}: {e}")
                        continue

            logger.info(f"Extraídos {len(datos_extraidos)} registros de países")
            return datos_extraidos

        except Exception as e:
            logger.error(f"Error leyendo Excel: {e}")
            raise

    def _convertir_a_numero(self, valor):
        """Convertir valor a número, manejando diferentes formatos"""
        if valor is None:
            return None

        if isinstance(valor, (int, float)):
            return int(valor) if isinstance(valor, float) and valor.is_integer() else valor

        if isinstance(valor, str):
            # Limpiar el string
            valor_limpio = valor.replace(',', '').replace('.', '').strip()
            try:
                return int(valor_limpio)
            except ValueError:
                try:
                    return float(valor)
                except ValueError:
                    return None

        return None

    def insertar_datos_en_db(self, datos):
        """Insertar los datos extraídos en la base de datos"""
        if not datos:
            logger.warning("No hay datos para insertar")
            return

        try:
            cursor = self.connection.cursor()

            # Filtrar: solo registros que tengan al menos uno de los 2 campos
            datos_filtrados = [
                d for d in datos
                if (d.get('viajeros_hoteles') is not None or d.get('pernoctaciones_hoteles') is not None)
            ]

            if not datos_filtrados:
                logger.warning("No hay registros válidos (viajeros/pernoctaciones) para insertar")
                return

            # Limpiar datos existentes para el mismo período (usando sólo periodos presentes en datos válidos)
            año_mes_pairs = set((d['año'], d['mes']) for d in datos_filtrados)

            for año, mes in año_mes_pairs:
                cursor.execute(
                    "DELETE FROM turismo_paises WHERE año = %s AND mes = %s",
                    (año, mes)
                )
                logger.info(f"Eliminados datos existentes para {mes}/{año}")

            # Preparar datos para inserción
            valores_insercion = []
            for dato in datos_filtrados:
                valores_insercion.append((
                    dato['año'],
                    dato['mes'],
                    dato['codigo_pais'],
                    dato['nombre_pais'],
                    dato['viajeros_hoteles'],
                    dato['pernoctaciones_hoteles']
                ))

            # Insertar datos
            query_insercion = """
                INSERT INTO turismo_paises 
                (año, mes, codigo_pais, nombre_pais, viajeros_hoteles, pernoctaciones_hoteles)
                VALUES %s
                ON CONFLICT (año, mes, codigo_pais) DO NOTHING
            """

            execute_values(cursor, query_insercion, valores_insercion)

            self.connection.commit()
            logger.info(f"Insertados {len(valores_insercion)} registros en la base de datos")

        except Exception as e:
            self.connection.rollback()
            logger.error(f"Error insertando datos: {e}")
            raise
        finally:
            cursor.close()

    def procesar_archivo(self):
        """Proceso principal: extraer datos del Excel e insertarlos en la DB"""
        try:
            self.conectar_db()
            datos = self.leer_excel_y_extraer_datos()
            self.insertar_datos_en_db(datos)
            logger.info("Proceso completado exitosamente")

        except Exception as e:
            logger.error(f"Error en el proceso: {e}")
            raise
        finally:
            if self.connection:
                self.connection.close()
                logger.info("Conexión a la base de datos cerrada")


def main():
    # Configuración de la base de datos
    db_config = {
        'host': 'localhost',  # Cambiar por tu host
        'database': 'nexus',  # Nombre de tu base de datos
        'user': 'maxim',  # Cambiar por tu usuario
        'password': 'diganDar',  # Cambiar por tu contraseña
        'port': 5432  # Puerto de PostgreSQL
    }

    # Archivo Excel a procesar
    dir_in = './nexus/data/ultimos_datos_turisticos'
    files = [fn for fn in os.listdir(dir_in) if '08_otros_mercados_' in fn and '.xls' in fn and '~' not in fn]

    excel_file = '08_otros_mercados_jun25.xlsx'
    for excel_file in files:
        excel_file_fname = os.path.join(dir_in, excel_file)

        # Crear el extractor y procesar
        extractor = TurismoDataExtractor(excel_file_fname, db_config)
        extractor.procesar_archivo()


if __name__ == "__main__":
    main()
